# 2025년도에 만든 교실 자리 뽑기 프로그램 V2
# 2025.11.03 - V2.6.5
import tkinter as tk
from tkinter import *
import random as r
from tkinter import messagebox, filedialog
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import range_boundaries
import os

# 좌석 배치 고정 상수
Title_Text = "🎓 교실 자리 배치 프로그램 🎓"
TOTAL_SEATS = 18
COLS = 6
ROWS = 3
MAX_REPEAT = 100  # 자동 배치 최대 반복 횟수
AUTO_RUN_DELAY_MS = 250  # 자동 배치 간격(ms) - 250밀리초

# 전역 변수
excluded = set()  # 제외할 번호
selected = set()  # 비활성화된 자리 번호
seat_buttons = []  # 자리 버튼들
is_seat_creation_phase = False  # 자리 생성 단계인지 여부
first_selected_seat = None  # 첫 번째 선택된 자리
current_seat_assignment = {}  # 현재 자리 배정 상태
current_scale = 1.0  # 현재 UI 크기 배율
update_window = None  # 업데이트 내용 편집 창
auto_run_active = False  # 자동 반복 실행 중 여부
speed_factor = 1.0      # 자동 반복 속도 배수(1.0=기본, 0.5=2배 빠름)
speed_key_press_count = 0  # 속도키('1') 누른 횟수
speed_key_press_times = []  # 속도키 입력 타임스탬프 기록
speed_boost_timeout = 1.0  # 1초 이내에 2번 눌러야 함
boost_count = 0         # 적용된 2배 속도 부스트 횟수(0=기본,1=x2,2=x4,3=x8)
current_theme = 'green'  # 현재 테마
theme_window = None  # 테마 설정 창

# 테마 색상 정의
THEMES = {
    'light': {
        'bg': 'white',
        'fg': 'black',
        'title_fg': '#F30000',
        'input_bg': 'white',
        'input_fg': 'black',
        'button_generate': '#d0d0d0',
        'button_assign': '#d0d0d0',
        'button_excel': '#d0d0d0',
        'button_fg': 'black',
        'seat_bg': 'lightblue',
        'seat_fg': 'black',
        'disabled_seat_bg': 'lightgray',
        'disabled_seat_fg': 'black',
        'selected_seat_bg': 'yellow',
        'selected_seat_fg': 'black',
        'blackboard_bg': '#d0d0d0',
        'blackboard_fg': 'black',
        'info_fg': '#666666',
        'countdown_fg': 'red'
    },
    'dark': {
        'bg': '#0d0d0d',
        'fg': '#e0e0e0',
        'title_fg': '#ff6b6b',
        'input_bg': '#e0e0e0',
        'input_fg': '#1a1a1a',
        'button_generate': '#1a1a1a',
        'button_assign': '#1a1a1a',
        'button_excel': '#1a1a1a',
        'button_fg': '#1a1a1a',
        'seat_bg': '#e0e0e0',
        'seat_fg': '#1a1a1a',
        'disabled_seat_bg': '#e0e0e0',
        'disabled_seat_fg': '#1a1a1a',
        'selected_seat_bg': '#e0e0e0',
        'selected_seat_fg': '#1a1a1a',
        'blackboard_bg': '#e0e0e0',
        'blackboard_fg': '#1a1a1a',
        'info_fg': '#e0e0e0',
        'countdown_fg': '#ff6b6b'
    },
    'blue': {
        'bg': '#e3f2fd',
        'fg': '#0d47a1',
        'title_fg': '#1565c0',
        'input_bg': '#ffffff',
        'input_fg': '#0d47a1',
        'button_generate': '#1976d2',
        'button_assign': '#0288d1',
        'button_excel': '#0277bd',
        'button_fg': 'black',
        'seat_bg': '#90caf9',
        'seat_fg': '#0d47a1',
        'disabled_seat_bg': '#bbdefb',
        'disabled_seat_fg': '#0d47a1',
        'selected_seat_bg': '#ffeb3b',
        'selected_seat_fg': '#000000',
        'blackboard_bg': '#0277bd',
        'blackboard_fg': 'black',
        'info_fg': '#1565c0',
        'countdown_fg': '#d32f2f'
    },
    'green': {
        'bg': '#e8f5e9',
        'fg': '#1b5e20',
        'title_fg': '#2e7d32',
        'input_bg': '#ffffff',
        'input_fg': '#1b5e20',
        'button_generate': '#43a047',
        'button_assign': '#66bb6a',
        'button_excel': '#4caf50',
        'button_fg': 'black',
        'seat_bg': '#a5d6a7',
        'seat_fg': '#1b5e20',
        'disabled_seat_bg': '#c8e6c9',
        'disabled_seat_fg': '#1b5e20',
        'selected_seat_bg': '#ffeb3b',
        'selected_seat_fg': '#000000',
        'blackboard_bg': '#4caf50',
        'blackboard_fg': 'black',
        'info_fg': '#2e7d32',
        'countdown_fg': '#d32f2f'
    },
    'purple': {
        'bg': '#f3e5f5',
        'fg': '#4a148c',
        'title_fg': '#6a1b9a',
        'input_bg': '#ffffff',
        'input_fg': '#4a148c',
        'button_generate': '#8e24aa',
        'button_assign': '#ab47bc',
        'button_excel': '#ba68c8',
        'button_fg': 'black',
        'seat_bg': '#ce93d8',
        'seat_fg': '#4a148c',
        'disabled_seat_bg': '#e1bee7',
        'disabled_seat_fg': '#4a148c',
        'selected_seat_bg': '#ffeb3b',
        'selected_seat_fg': '#000000',
        'blackboard_bg': '#ba68c8',
        'blackboard_fg': 'black',
        'info_fg': '#6a1b9a',
        'countdown_fg': '#d32f2f'
    }
}

def zoom_in(event=None):
    """UI 확대 (Command + '+' 또는 Command + '=')"""
    global current_scale
    if current_scale < 1.8:  # 최대 1.8배까지만 확대 (180%)
        current_scale += 0.1
        print(f"확대: {current_scale:.1f} ({int(current_scale * 100)}%)")  # 디버깅용
        apply_zoom()

def zoom_out(event=None):
    """UI 축소 (Command + '-')"""
    global current_scale
    if current_scale > 1.0:  # 최소 1.0배까지만 축소 (100%)
        current_scale -= 0.1
        print(f"축소: {current_scale:.1f} ({int(current_scale * 100)}%)")  # 디버깅용
        apply_zoom()

def zoom_reset(event=None):
    """UI 원래 크기로 복원 (Command + '0')"""
    global current_scale
    current_scale = 1.0
    print(f"원래 크기: {current_scale:.1f} ({int(current_scale * 100)}%)")  # 디버깅용
    apply_zoom()

def apply_zoom():
    """현재 배율을 모든 UI 요소에 적용"""
    try:
        # 폰트 크기 조정
        base_font_size = 12
        base_title_font_size = 20
        base_countdown_font_size = 40
        
        # 입력 필드 폰트 크기 조정
        new_font_size = int(base_font_size * current_scale)
        new_title_font_size = int(base_title_font_size * current_scale)
        new_countdown_font_size = int(base_countdown_font_size * current_scale)
        
        # 라벨 폰트 크기 조정 (존재하는 경우에만) - Frame으로 감싸져 있으므로 주석 처리
        # if 'label_grade' in globals() and label_grade.winfo_exists():
        #     label_grade.config(font=('맑은 고딕', new_font_size, 'bold'))
        # if 'label_group' in globals() and label_group.winfo_exists():
        #     label_group.config(font=('맑은 고딕', new_font_size, 'bold'))
        # if 'label_students' in globals() and label_students.winfo_exists():
        #     label_students.config(font=('맑은 고딕', new_font_size, 'bold'))
        # if 'label_teacher' in globals() and label_teacher.winfo_exists():
        #     label_teacher.config(font=('맑은 고딕', new_font_size, 'bold'))
        if 'label_exclude' in globals() and label_exclude.winfo_exists():
            label_exclude.config(font=('맑은 고딕', new_font_size, 'bold'))
        if 'label_repeat' in globals() and label_repeat.winfo_exists():
            label_repeat.config(font=('맑은 고딕', new_font_size, 'bold'))
        
        # 입력 필드 폰트 크기 조정 (존재하는 경우에만)
        if 'entry_grade' in globals() and entry_grade.winfo_exists():
            entry_grade.config(font=('맑은 고딕', new_font_size))
        if 'entry_group' in globals() and entry_group.winfo_exists():
            entry_group.config(font=('맑은 고딕', new_font_size))
        if 'entry_students' in globals() and entry_students.winfo_exists():
            entry_students.config(font=('맑은 고딕', new_font_size))
        if 'entry_teacher' in globals() and entry_teacher.winfo_exists():
            entry_teacher.config(font=('맑은 고딕', new_font_size))
        if 'entry_exclude' in globals() and entry_exclude.winfo_exists():
            entry_exclude.config(font=('맑은 고딕', new_font_size))
        if 'entry_repeat' in globals() and entry_repeat.winfo_exists():
            entry_repeat.config(font=('맑은 고딕', new_font_size))
        
        # 버튼 폰트 크기 조정 (존재하는 경우에만)
        if 'btn_generate_candidates' in globals() and btn_generate_candidates.winfo_exists():
            btn_generate_candidates.config(font=('맑은 고딕', int(11 * current_scale), 'bold'))
        if 'btn_generate_seats' in globals() and btn_generate_seats.winfo_exists():
            btn_generate_seats.config(font=('맑은 고딕', int(11 * current_scale), 'bold'))
        if 'btn_create_excel' in globals() and btn_create_excel.winfo_exists():
            btn_create_excel.config(font=('맑은 고딕', int(11 * current_scale), 'bold'))
        
        # 설명 라벨 폰트 크기 조정 (존재하는 경우에만)
        if 'info_label' in globals() and info_label.winfo_exists():
            info_label.config(font=('맑은 고딕', new_title_font_size))
        
        # 칠판 라벨 폰트 크기 조정 (존재하는 경우에만)
        if 'blackboard_label' in globals() and blackboard_label.winfo_exists():
            blackboard_label.config(font=('맑은 고딕', int(11 * current_scale), 'bold'))
        
        # 카운트다운 라벨 폰트 크기 조정 (존재하는 경우에만)
        if 'countdown_label' in globals() and countdown_label.winfo_exists():
            countdown_label.config(font=('맑은 고딕', new_countdown_font_size, 'bold'))
        
        # 자리 버튼들 폰트 크기 조정 (이미 생성된 경우)
        for row_buttons in seat_buttons:
            for btn in row_buttons:
                if btn.winfo_exists():
                    btn.config(font=('맑은 고딕', int(12 * current_scale)))
        
        # 상태 표시 업데이트
        if 'zoom_status_label' in globals() and zoom_status_label.winfo_exists():
            zoom_status_label.config(text=f"확대/축소: {int(current_scale * 100)}%")
            
    except Exception as e:
        print(f"확대/축소 적용 중 오류: {e}")

def apply_theme(theme_name):
    """선택한 테마를 전체 UI에 적용"""
    global current_theme
    
    # 이미 같은 테마가 적용되어 있으면 중복 적용하지 않음
    if current_theme == theme_name:
        return
    
    current_theme = theme_name
    theme = THEMES[theme_name]
    
    try:
        # 메인 윈도우 및 컨테이너
        root.config(bg=theme['bg'])
        main_container.config(bg=theme['bg'])
        
        # 타이틀
        title_label.config(bg=theme['bg'], fg=theme['title_fg'])
        
        # 입력 프레임
        input_frame.config(bg=theme['bg'])
        
        # 입력 필드 라벨 프레임들
        label_grade_frame.config(bg=theme['bg'])
        label_group_frame.config(bg=theme['bg'])
        label_students_frame.config(bg=theme['bg'])
        label_teacher_frame.config(bg=theme['bg'])
        
        # 라벨들
        for widget in [label_grade_frame, label_group_frame, label_students_frame, label_teacher_frame]:
            for child in widget.winfo_children():
                if isinstance(child, Label):
                    if child['fg'] == 'red':  # * 표시는 빨간색 유지
                        child.config(bg=theme['bg'], fg='red')
                    else:
                        child.config(bg=theme['bg'], fg=theme['fg'])
        
        label_exclude.config(bg=theme['bg'], fg=theme['fg'])
        label_repeat.config(bg=theme['bg'], fg=theme['fg'])
        
        # 입력 필드들
        for entry in [entry_grade, entry_group, entry_students, entry_teacher, entry_exclude, entry_repeat]:
            entry.config(bg=theme['input_bg'], fg=theme['input_fg'], insertbackground=theme['input_fg'])
        
        # 설명 라벨
        info_label.config(bg=theme['bg'], fg=theme['info_fg'])
        
        # 버튼 프레임
        btn_frame.config(bg=theme['bg'])
        
        # 버튼들
        btn_generate_candidates.config(bg=theme['button_generate'], fg=theme['button_fg'], 
                                       highlightbackground=theme['button_generate'], 
                                       activebackground=theme['button_generate'])
        btn_generate_seats.config(bg=theme['button_assign'], fg=theme['button_fg'],
                                 highlightbackground=theme['button_assign'],
                                 activebackground=theme['button_assign'])
        btn_create_excel.config(bg=theme['button_excel'], fg=theme['button_fg'],
                               highlightbackground=theme['button_excel'],
                               activebackground=theme['button_excel'])
        
        # 칠판
        blackboard_label.config(bg=theme['blackboard_bg'], fg=theme['blackboard_fg'],
                               highlightbackground=theme['blackboard_bg'],
                               activebackground=theme['blackboard_bg'])
        
        # 자리 배치 프레임
        frame.config(bg=theme['bg'])
        
        # 카운트다운
        countdown_label.config(bg=theme['bg'], fg=theme['countdown_fg'])
        
        # 업데이트 버튼 프레임
        update_button_frame.config(bg=theme['bg'])
        update_content_btn.config(bg=theme['button_generate'], fg=theme['button_fg'],
                                 highlightbackground=theme['button_generate'],
                                 activebackground=theme['button_generate'])
        
        # 테마 버튼 (추가될 예정)
        if 'theme_button_frame' in globals() and theme_button_frame.winfo_exists():
            theme_button_frame.config(bg=theme['bg'])
            theme_btn.config(bg=theme['button_excel'], fg=theme['button_fg'],
                           highlightbackground=theme['button_excel'],
                           activebackground=theme['button_excel'])
        
        # 상태 라벨들
        zoom_status_label.config(bg=theme['bg'], fg=theme['fg'])
        shortcut_label.config(bg=theme['bg'], fg=theme['info_fg'])
        
        # 자리 버튼들 (생성된 경우)
        for row_buttons in seat_buttons:
            for btn in row_buttons:
                if btn.winfo_exists():
                    current_text = btn['text']
                    current_bg = btn['bg']
                    
                    if current_text == 'X':  # 비활성화된 자리
                        btn.config(bg=theme['disabled_seat_bg'], fg=theme['disabled_seat_fg'])
                    elif current_text == '':  # 빈 자리 (애니메이션 전)
                        btn.config(bg=theme['seat_bg'], fg=theme['seat_fg'])
                    else:  # 학생 이름이 있는 자리
                        # 선택된 자리인지 확인 (이전 테마의 selected_seat_bg 색상이거나 현재 테마의 색상)
                        is_selected = False
                        for theme_name in THEMES:
                            if current_bg == THEMES[theme_name]['selected_seat_bg']:
                                is_selected = True
                                break
                        
                        if is_selected:
                            btn.config(bg=theme['selected_seat_bg'], fg=theme['selected_seat_fg'])
                        else:
                            btn.config(bg=theme['seat_bg'], fg=theme['seat_fg'])
        
        print(f"테마 적용됨: {theme_name}")
    except Exception as e:
        print(f"테마 적용 중 오류: {e}")

def open_theme_selector():
    """테마 선택 창을 엽니다"""
    global theme_window
    
    # 이미 열려있다면 포커스만 이동
    if theme_window and theme_window.winfo_exists():
        theme_window.lift()
        theme_window.focus_force()
        return
    
    # 새 창 생성
    theme_window = Toplevel(root)
    theme_window.title("테마 설정")
    theme_window.geometry("500x400")
    theme_window.config(bg=THEMES[current_theme]['bg'])
    theme_window.resizable(False, False)
    
    # 창이 닫힐 때 전역 변수 정리
    def on_closing():
        global theme_window
        try:
            theme_window.destroy()
        except:
            pass
        theme_window = None
    
    theme_window.protocol("WM_DELETE_WINDOW", on_closing)
    
    # 메인 프레임
    main_frame = Frame(theme_window, bg=THEMES[current_theme]['bg'])
    main_frame.pack(expand=True, fill='both', padx=20, pady=20)
    
    # 제목
    title = Label(main_frame, text="🎨 테마 선택", 
                 font=('맑은 고딕', 18, 'bold'), 
                 bg=THEMES[current_theme]['bg'], 
                 fg=THEMES[current_theme]['title_fg'])
    title.pack(pady=(0, 10))
    
    # 설명
    info = Label(main_frame, text="원하는 테마를 선택해주세요", 
                font=('맑은 고딕', 10), 
                bg=THEMES[current_theme]['bg'], 
                fg=THEMES[current_theme]['info_fg'])
    info.pack(pady=(0, 20))
    
    # 테마 버튼들
    themes_info = [
        ('light', '☀️ 라이트 모드', '밝고 깔끔한 기본 테마'),
        ('dark', '🌙 다크 모드', '어두운 배경의 눈이 편한 테마'),
        ('blue', '💙 블루 테마', '시원한 파란색 테마'),
        ('green', '💚 그린 테마', '자연스러운 녹색 테마'),
        ('purple', '💜 퍼플 테마', '고급스러운 보라색 테마')
    ]
    
    for theme_key, theme_name, theme_desc in themes_info:
        btn_frame = Frame(main_frame, bg=THEMES[current_theme]['bg'])
        btn_frame.pack(pady=5, fill='x')
        
        def make_select_theme(tk=theme_key):
            def select():
                apply_theme(tk)
                on_closing()
            return select
        
        theme_btn = Button(btn_frame, text=theme_name, 
                          command=make_select_theme(),
                          font=('맑은 고딕', 12, 'bold'), 
                          bg=THEMES[theme_key]['button_generate'], 
                          fg=THEMES[theme_key]['button_fg'],
                          relief='raised', bd=2, width=20, height=2)
        theme_btn.pack(side='left', padx=5)
        
        desc_label = Label(btn_frame, text=theme_desc, 
                          font=('맑은 고딕', 9), 
                          bg=THEMES[current_theme]['bg'], 
                          fg=THEMES[current_theme]['info_fg'])
        desc_label.pack(side='left', padx=10)
        
        # 현재 선택된 테마 표시
        if theme_key == current_theme:
            check_label = Label(btn_frame, text="✓", 
                              font=('맑은 고딕', 14, 'bold'), 
                              bg=THEMES[current_theme]['bg'], 
                              fg=THEMES[current_theme]['countdown_fg'])
            check_label.pack(side='left')



def toggle_exclude(num, button):
    theme = THEMES[current_theme]
    if num in excluded:
        excluded.remove(num)
        button.config(bg=theme['seat_bg'], text='')
    else:
        excluded.add(num)
        button.config(bg=theme['disabled_seat_bg'], text='X')

def add_excluded_numbers():
    try:
        # 기존 제외 목록 초기화
        excluded.clear()
        
        # 입력된 번호들을 처리
        numbers = entry_exclude.get().strip()
        # 전체 학생 수 (개별 번호 유효성 검사에 사용)
        total_students = None
        try:
            total_students = int(entry_students.get())
        except Exception:
            pass
        if numbers:
            # 쉼표로 구분된 번호들을 처리
            for num in numbers.split(','):
                num = num.strip()
                if num:
                    num = int(num)
                    if num <= 0:
                        messagebox.showerror("오류", "1 이상의 숫자만 입력 가능합니다!")
                        return False
                    if total_students is not None and num > total_students:
                        messagebox.showerror("오류", "제외할 번호가 학생 수 범위를 초과했습니다!")
                        return False
                    excluded.add(num)
        
        # 제외된 번호가 전체 학생 수보다 많으면 경고
        if len(excluded) > int(entry_students.get()):
            messagebox.showerror("오류", "제외할 번호가 전체 학생 수보다 많습니다!")
            return False
            
        return True
    except ValueError:
        messagebox.showerror("오류", "올바른 숫자를 입력해주세요!")
        return False

def generate_candidate_buttons():
    global seat_buttons, selected, is_seat_creation_phase
    for widget in frame.winfo_children():
        widget.destroy()
    seat_buttons = []
    selected = set()
    is_seat_creation_phase = True
    
    # 자리 배치 버튼 활성화
    btn_generate_seats.config(state='normal')
    


    try:
        nums = int(entry_students.get())
        if nums <= 0:
            messagebox.showerror("오류", "올바른 학생 수를 입력해주세요!")
            return
        if nums > TOTAL_SEATS:
            messagebox.showerror("오류", f"학생 수는 {TOTAL_SEATS}명 이하로만 입력 가능합니다!")
            return
    except ValueError:
        messagebox.showerror("오류", "올바른 학생 수를 입력해주세요!")
        return

    total_seats = TOTAL_SEATS
    cols = COLS
    rows = ROWS
    
    theme = THEMES[current_theme]

    for i in range(rows):
        row_buttons = []
        for j in range(cols):
            idx = i * cols + j + 1
            if idx > total_seats:
                break

            btn = Button(frame, text='', width=8, height=3, font=('맑은 고딕', int(12 * current_scale)),
                         bg=theme['seat_bg'], fg=theme['seat_fg'], command=lambda i=i, j=j: select_seat(i, j))

            # 그룹 간 간격 조정 (2개씩 붙이고 그룹 사이 넓게)
            if j % 2 == 0:
                padx_val = (0, 2)  # 왼쪽 끝이면 오른쪽 약간 간격
            else:
                padx_val = (0, 10)  # 짝꿍 오른쪽 끝이면 그룹 사이 간격 넓게

            btn.grid(row=i, column=j, padx=padx_val, pady=5)
            row_buttons.append(btn)
        seat_buttons.append(row_buttons)


def select_seat(i, j):
    global selected, first_selected_seat
    idx = i * len(seat_buttons[0]) + j + 1
    theme = THEMES[current_theme]
    
    # 자리 생성 단계에서는 자리 비활성화
    if is_seat_creation_phase:
        if idx in selected:
            selected.remove(idx)
            seat_buttons[i][j].config(bg=theme['seat_bg'], fg=theme['seat_fg'], text='')
        else:
            selected.add(idx)
            seat_buttons[i][j].config(bg=theme['disabled_seat_bg'], fg=theme['disabled_seat_fg'], text='X')
    # 자리 배치 단계에서는 자리 교환
    else:
        # 비활성화된 자리(X)는 선택할 수 없음
        if seat_buttons[i][j]['text'] == 'X':
            return
            
        if first_selected_seat is None:
            first_selected_seat = (i, j)
            seat_buttons[i][j].config(bg=theme['selected_seat_bg'], fg=theme['selected_seat_fg'])
        else:
            # 두 번째 자리 선택 시 교환
            i1, j1 = first_selected_seat
            # 첫 번째 선택된 자리의 텍스트와 배경색 저장
            temp_text = seat_buttons[i1][j1]['text']
            temp_bg = seat_buttons[i1][j1]['bg']
            
            # 두 자리의 텍스트와 배경색 교환
            seat_buttons[i1][j1].config(text=seat_buttons[i][j]['text'], bg=theme['seat_bg'], fg=theme['seat_fg'])
            seat_buttons[i][j].config(text=temp_text, bg=theme['seat_bg'], fg=theme['seat_fg'])
            
            # 자리 배정 상태 업데이트
            if temp_text and seat_buttons[i][j]['text']:
                current_seat_assignment[temp_text] = (i, j)
                current_seat_assignment[seat_buttons[i][j]['text']] = (i1, j1)
            
            # 첫 번째 선택 초기화
            first_selected_seat = None

def animate_seat_shuffle(buttons_data, iteration=0, max_iterations=8):
    """좌석들이 랜덤하게 위치를 바꾸는 셔플 애니메이션"""
    theme = THEMES[current_theme]
    
    if iteration < max_iterations:
        # 활성화된 좌석들의 텍스트를 랜덤하게 섞기
        active_buttons = [(btn, text) for btn, text in buttons_data if text != 'X']
        
        if len(active_buttons) > 1:
            # 텍스트만 추출
            texts = [text for _, text in active_buttons]
            # 랜덤하게 섞기
            shuffled_texts = texts.copy()
            r.shuffle(shuffled_texts)
            
            # 섞인 텍스트를 버튼에 적용
            for (btn, _), new_text in zip(active_buttons, shuffled_texts):
                if btn.winfo_exists():
                    btn.config(text=new_text)
        
        # 점점 느려지는 딜레이 (50ms -> 200ms)
        delay = 50 + (iteration * 20)
        root.after(delay, lambda: animate_seat_shuffle(buttons_data, iteration + 1, max_iterations))
    # else: 셔플 완료 - 최종 위치는 이후 슬롯머신 애니메이션에서 결정

def animate_slot_machine(btn, final_number, available_students, iteration=0, max_iterations=15, show_highlight=True):
    """슬롯머신처럼 숫자가 돌아가는 애니메이션"""
    theme = THEMES[current_theme]
    
    # 버튼이 존재하는지 확인
    if not btn.winfo_exists():
        return
    
    if iteration < max_iterations:
        # 랜덤 숫자 표시 (점점 느려지는 효과)
        random_num = r.choice(available_students)
        btn.config(text=str(random_num))
        
        # 점점 느려지는 딜레이 (30ms -> 150ms)
        delay = 30 + (iteration * 8)
        root.after(delay, lambda: animate_slot_machine(btn, final_number, available_students, iteration + 1, max_iterations, show_highlight))
    else:
        # 최종 숫자 표시
        btn.config(text=str(final_number))
        if show_highlight:
            # 깜빡이는 효과 (1회 실행 또는 마지막 반복일 때만)
            btn.config(bg=theme['selected_seat_bg'], fg=theme['selected_seat_fg'])
            root.after(200, lambda: btn.config(bg=theme['seat_bg'], fg=theme['seat_fg']) if btn.winfo_exists() else None)
        else:
            # 하이라이트 없이 바로 표시
            btn.config(bg=theme['seat_bg'], fg=theme['seat_fg'])

def animate_slot_machine_quick(btn, final_number, available_students, iteration=0, max_iterations=5):
    """빠른 슬롯머신 애니메이션 (자동 반복용)"""
    theme = THEMES[current_theme]
    
    # 버튼이 존재하는지 확인
    if not btn.winfo_exists():
        return
    
    if iteration < max_iterations:
        # 랜덤 숫자 표시 (빠른 속도)
        random_num = r.choice(available_students)
        btn.config(text=str(random_num))
        
        # 빠른 딜레이 (20ms -> 50ms)
        delay = 20 + (iteration * 6)
        root.after(delay, lambda: animate_slot_machine_quick(btn, final_number, available_students, iteration + 1, max_iterations))
    else:
        # 최종 숫자 표시 (하이라이트 없음)
        btn.config(text=str(final_number), bg=theme['seat_bg'], fg=theme['seat_fg'])

def generate_seats_with_animation(is_last_iteration=True):
    """애니메이션과 함께 자리를 배치하는 함수"""
    global seat_buttons, selected, is_seat_creation_phase, first_selected_seat, current_seat_assignment
    for widget in frame.winfo_children():
        widget.destroy()
    seat_buttons = []
    is_seat_creation_phase = False
    first_selected_seat = None
    current_seat_assignment.clear()

    try:
        nums = int(entry_students.get())
        if nums <= 0:
            messagebox.showerror("오류", "올바른 학생 수를 입력해주세요!")
            return
        if nums > TOTAL_SEATS:
            messagebox.showerror("오류", f"학생 수는 {TOTAL_SEATS}명 이하로만 입력 가능합니다!")
            return
    except ValueError:
        messagebox.showerror("오류", "올바른 학생 수를 입력해주세요!")
        return

    if not add_excluded_numbers():
        return

    total_seats = TOTAL_SEATS
    cols = COLS
    rows = ROWS
    theme = THEMES[current_theme]

    available_students = [i for i in range(1, nums + 1) if i not in excluded]
    r.shuffle(available_students)

    active_seats = total_seats - len(selected)
    if active_seats != len(available_students):
        messagebox.showerror("오류", "활성화된 자리 수와 배정할 학생 수가 일치하지 않습니다!")
        return

    # 먼저 모든 버튼을 생성 (애니메이션용)
    student_idx = 0
    animation_data = []  # (버튼, 최종번호) 저장
    shuffle_data = []  # 셔플 애니메이션용 (버튼, 임시텍스트)
    
    for i in range(rows):
        row_buttons = []
        for j in range(cols):
            idx = i * cols + j + 1
            if idx > total_seats:
                break
            
            if idx in selected:
                btn = Button(frame, text='X', width=8, height=3, font=('맑은 고딕', int(12 * current_scale)),
                             bg=theme['disabled_seat_bg'], fg=theme['disabled_seat_fg'], state='disabled')
                shuffle_data.append((btn, 'X'))
            elif student_idx < len(available_students):
                student = available_students[student_idx]
                student_idx += 1
                # 초기에는 임시 번호를 표시 (셔플용)
                temp_number = str(student)
                btn = Button(frame, text=temp_number, width=8, height=3, font=('맑은 고딕', int(12 * current_scale)),
                             bg=theme['seat_bg'], fg=theme['seat_fg'], command=lambda i=i, j=j: select_seat(i, j))
                current_seat_assignment[str(student)] = (i, j)
                animation_data.append((btn, student))
                shuffle_data.append((btn, temp_number))
            else:
                btn = Button(frame, text='', width=8, height=3, font=('맑은 고딕', int(12 * current_scale)),
                             bg=theme['disabled_seat_bg'], fg=theme['disabled_seat_fg'], state='disabled')
                shuffle_data.append((btn, ''))
            
            if j % 2 == 0:
                padx_val = (0, 2)
            else:
                padx_val = (0, 10)
            btn.grid(row=i, column=j, padx=padx_val, pady=5)
            row_buttons.append(btn)
        seat_buttons.append(row_buttons)
    
    # 1단계: 좌석 셔플 애니메이션 (약 1.2초)
    shuffle_duration = 1200  # 8회 * (50ms + 20ms*7.5 평균) ≈ 1200ms
    animate_seat_shuffle(shuffle_data, iteration=0, max_iterations=8)
    
    # 2단계: 셔플 완료 후 슬롯머신 애니메이션 시작
    def start_slot_machine():
        # 순차적으로 애니메이션 시작 (각 버튼마다 약간씩 딜레이)
        for idx, (btn, final_num) in enumerate(animation_data):
            delay = idx * 50  # 50ms씩 순차적으로 시작
            root.after(delay, lambda b=btn, f=final_num: animate_slot_machine(b, f, available_students, show_highlight=is_last_iteration))
    
    root.after(shuffle_duration, start_slot_machine)
    
    # 모든 애니메이션이 끝난 후 버튼 활성화
    total_animation_time = shuffle_duration + len(animation_data) * 50 + 15 * 100  # 셔플 + 슬롯머신 총 시간
    def enable_buttons():
        for row_buttons in seat_buttons:
            for btn in row_buttons:
                if btn['text'] != 'X':
                    btn.config(state='normal')
    
    root.after(total_animation_time, enable_buttons)

def generate_seats_with_quick_animation():
    """빠른 애니메이션과 함께 자리를 배치하는 함수 (자동 반복용)"""
    global seat_buttons, selected, is_seat_creation_phase, first_selected_seat, current_seat_assignment
    for widget in frame.winfo_children():
        widget.destroy()
    seat_buttons = []
    is_seat_creation_phase = False
    first_selected_seat = None
    current_seat_assignment.clear()

    try:
        nums = int(entry_students.get())
        if nums <= 0:
            return
        if nums > TOTAL_SEATS:
            return
    except ValueError:
        return

    if not add_excluded_numbers():
        return

    total_seats = TOTAL_SEATS
    cols = COLS
    rows = ROWS
    theme = THEMES[current_theme]

    available_students = [i for i in range(1, nums + 1) if i not in excluded]
    r.shuffle(available_students)

    active_seats = total_seats - len(selected)
    if active_seats != len(available_students):
        return

    # 모든 버튼 생성
    student_idx = 0
    animation_data = []
    
    for i in range(rows):
        row_buttons = []
        for j in range(cols):
            idx = i * cols + j + 1
            if idx > total_seats:
                break
            
            if idx in selected:
                btn = Button(frame, text='X', width=8, height=3, font=('맑은 고딕', int(12 * current_scale)),
                             bg=theme['disabled_seat_bg'], fg=theme['disabled_seat_fg'], state='disabled')
            elif student_idx < len(available_students):
                student = available_students[student_idx]
                student_idx += 1
                btn = Button(frame, text='?', width=8, height=3, font=('맑은 고딕', int(12 * current_scale)),
                             bg=theme['seat_bg'], fg=theme['seat_fg'], command=lambda i=i, j=j: select_seat(i, j), state='disabled')
                current_seat_assignment[str(student)] = (i, j)
                animation_data.append((btn, student))
            else:
                btn = Button(frame, text='', width=8, height=3, font=('맑은 고딕', int(12 * current_scale)),
                             bg=theme['disabled_seat_bg'], fg=theme['disabled_seat_fg'], state='disabled')
            
            if j % 2 == 0:
                padx_val = (0, 2)
            else:
                padx_val = (0, 10)
            btn.grid(row=i, column=j, padx=padx_val, pady=5)
            row_buttons.append(btn)
        seat_buttons.append(row_buttons)
    
    # 빠른 애니메이션 시작 (동시 시작)
    for btn, final_num in animation_data:
        animate_slot_machine_quick(btn, final_num, available_students)

def generate_seats():
    """애니메이션 없이 즉시 자리를 배치하는 함수 (자동 반복용)"""
    global seat_buttons, selected, is_seat_creation_phase, first_selected_seat, current_seat_assignment
    for widget in frame.winfo_children():
        widget.destroy()
    seat_buttons = []
    is_seat_creation_phase = False
    first_selected_seat = None
    current_seat_assignment.clear()

    try:
        nums = int(entry_students.get())
        if nums <= 0:
            messagebox.showerror("오류", "올바른 학생 수를 입력해주세요!")
            return
        if nums > TOTAL_SEATS:
            messagebox.showerror("오류", f"학생 수는 {TOTAL_SEATS}명 이하로만 입력 가능합니다!")
            return
    except ValueError:
        messagebox.showerror("오류", "올바른 학생 수를 입력해주세요!")
        return

    if not add_excluded_numbers():
        return

    total_seats = TOTAL_SEATS
    cols = COLS
    rows = ROWS
    theme = THEMES[current_theme]

    available_students = [i for i in range(1, nums + 1) if i not in excluded]
    r.shuffle(available_students)

    active_seats = total_seats - len(selected)
    if active_seats != len(available_students):
        messagebox.showerror("오류", "활성화된 자리 수와 배정할 학생 수가 일치하지 않습니다!")
        return

    student_idx = 0
    for i in range(rows):
        row_buttons = []
        for j in range(cols):
            idx = i * cols + j + 1
            if idx > total_seats:
                break
            
            if idx in selected:
                btn = Button(frame, text='X', width=8, height=3, font=('맑은 고딕', int(12 * current_scale)),
                             bg=theme['disabled_seat_bg'], fg=theme['disabled_seat_fg'], state='disabled')
            elif student_idx < len(available_students):
                student = available_students[student_idx]
                student_idx += 1
                btn = Button(frame, text=str(student), width=8, height=3, font=('맑은 고딕', int(12 * current_scale)),
                             bg=theme['seat_bg'], fg=theme['seat_fg'], command=lambda i=i, j=j: select_seat(i, j))
                current_seat_assignment[str(student)] = (i, j)
            else:
                btn = Button(frame, text='', width=8, height=3, font=('맑은 고딕', int(12 * current_scale)),
                             bg=theme['disabled_seat_bg'], fg=theme['disabled_seat_fg'], state='disabled')
            
            if j % 2 == 0:
                padx_val = (0, 2)
            else:
                padx_val = (0, 10)
            btn.grid(row=i, column=j, padx=padx_val, pady=5)
            row_buttons.append(btn)
        seat_buttons.append(row_buttons)
    
    for row_buttons in seat_buttons:
        for btn in row_buttons:
            if btn['text'] != 'X':
                btn.config(state='normal')

def create_excel_file():
    # 엑셀 파일 생성 함수
    try:
        # 입력값 검증
        grade = int(entry_grade.get())
        group = int(entry_group.get())
        n = int(entry_students.get())
        teacher = entry_teacher.get().strip()
        
        if not teacher:
            messagebox.showerror("오류", "담임선생님 성함을 입력해주세요!")
            return
            
        # 자리 생성 단계를 거치지 않았으면 경고
        if not seat_buttons:
            messagebox.showerror("오류", "먼저 자리 생성을 완료해주세요!")
            return
            
        if not current_seat_assignment:
            messagebox.showerror("오류", "먼저 자리 배치를 완료해주세요!")
            return
            
    except ValueError:
        messagebox.showerror("오류", "올바른 숫자를 입력해주세요!")
        return

    # 바탕화면 경로 설정
    desktop_path = os.path.expanduser("~/Desktop")
    default_filename = f"{grade}학년{group}반_좌석배정표.xlsx"
    default_path = os.path.join(desktop_path, default_filename)

    # 파일 저장 위치 선택
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        title="엑셀 파일 저장",
        initialdir=desktop_path,
        initialfile=default_filename
    )
    
    if not file_path:
        return

    # 엑셀 파일 생성
    xlsx = Workbook()
    x1 = xlsx.active

    # === 열 너비 조정 ===
    x1.column_dimensions['E'].width = 12
    x1.column_dimensions['H'].width = 12

    # === 인쇄 설정 추가 ===
    x1.page_setup.paperSize = x1.PAPERSIZE_A4
    x1.page_setup.orientation = 'landscape'
    x1.page_margins.left = 1.0
    x1.page_margins.right = 1.0
    x1.page_margins.top = 1.0
    x1.page_margins.bottom = 1.0
    x1.page_margins.header = 0.5
    x1.page_margins.footer = 0.5
    # =====================

    # 폰트 스타일 정의
    Title_font = Font(name='Pretendard', size=24, bold=True)
    Pretendard = Font(name='Pretendard', size=12, bold=True)

    # 테두리 스타일 정의
    Thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    thickleft_Thin_border = Border(
        left=Side(style='thick'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    thickright_Thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thick'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    Thick_border = Side(style='thick')
    No_border = Side(style=None)

    # 전체 범위 지정
    min_row, max_row = 1, 26  # 행
    min_col, max_col = 1, 12  # 열

    # 각 셀에 대해 위치에 따라 테두리 적용
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = x1.cell(row=row, column=col)

            # 각 방향 테두리 조건부 설정
            if row == min_row:
                top = Thick_border
                bottom = No_border
            elif row == max_row:
                top = No_border
                bottom = Thick_border
            else:
                top = No_border
                bottom = No_border

            if col == min_col:
                left = Thick_border
                right = No_border
            elif col == max_col:
                left = No_border
                right = Thick_border
            else:
                left = No_border
                right = No_border

            cell.border = Border(top=top, bottom=bottom, left=left, right=right)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 기본 병합
    x1.merge_cells('B2:K3')  # 좌석 배정표
    x1.merge_cells('B22:C22')  # 학반
    x1.merge_cells('B23:C24')  # 담임선생님
    x1.merge_cells('E23:H24')  # 칠판
    x1.merge_cells('A5:A19')  # 왼쪽 사물함
    x1.merge_cells('L5:L19')  # 오른쪽 사물함

    # 기본 테두리
    for row in x1['B2:K3']:  # 좌석 배치표(타이틀)
        for cell in row:
            cell.border = Thin_border

    for row in x1['B22:C22']:  # 학반
        for cell in row:
            cell.border = Thin_border

    for row in x1['B23:C24']:  # 담임 선생님
        for cell in row:
            cell.border = Thin_border

    for row in x1['E23:H24']:  # 칠판
        for cell in row:
            cell.border = Thin_border
    
    for row in x1['A5:A19']:  # 왼쪽 사물함
        for cell in row:
            cell.border = thickleft_Thin_border
    
    for row in x1['L5:L19']:  # 오른쪽 사물함
        for cell in row:
            cell.border = thickright_Thin_border

    # 기본 데이터 입력
    x1['B2'] = "좌석 배정표"
    x1['B2'].font = Title_font

    x1['E23'] = "칠판"
    x1['E23'].font = Pretendard

    # 자리표 배치 좌표 (기존 코드와 동일)
    seat_positions = [
        ('J15:J16', 15, 10),
        ('I15:I16', 15, 9),
        ('G15:G16', 15, 7),
        ('F15:F16', 15, 6),
        ('D15:D16', 15, 4),
        ('C15:C16', 15, 3),
        ('J11:J12', 11, 10),
        ('I11:I12', 11, 9),
        ('G11:G12', 11, 7),
        ('F11:F12', 11, 6),
        ('D11:D12', 11, 4),
        ('C11:C12', 11, 3),
        ('J7:J8', 7, 10),
        ('I7:I8', 7, 9),
        ('G7:G8', 7, 7),
        ('F7:F8', 7, 6),
        ('D7:D8', 7, 4),
        ('C7:C8', 7, 3)
    ]

    # 학생 리스트 준비 (제외 번호 제외)
    # students = [str(i) for i in range(1, n + 1) if i not in excluded]
    # student_idx = 0
    for idx, (merge_range, row, col) in enumerate(seat_positions):
        gui_row = idx // 6
        gui_col = idx % 6
        if seat_buttons and seat_buttons[gui_row][gui_col]['text'] == 'X':
            continue
        button_text = seat_buttons[gui_row][gui_col]['text']
        if button_text and button_text != 'X':
            x1.merge_cells(merge_range)
            x1.cell(row=row, column=col).value = button_text
            set_border_to_merged_range(x1, merge_range, Thin_border)
        # 학생이 없으면 빈 칸(아무것도 안함)

    # 추가 데이터 입력
    x1['B22'] = f'{grade}-{group}'
    x1['B22'].font = Pretendard

    x1['B23'] = teacher
    x1['B23'].font = Pretendard

    x1['A5'] = "사물함"
    x1['A5'].font = Pretendard
    
    x1['L5'] = "사물함"
    x1['L5'].font = Pretendard

    # 엑셀 파일 저장
    try:
        xlsx.save(file_path)
        messagebox.showinfo("성공", f"엑셀 파일이 성공적으로 저장되었습니다!\n저장 위치: {file_path}")
    except Exception as e:
        messagebox.showerror("오류", f"파일 저장 중 오류가 발생했습니다: {str(e)}")

def can_assign_seats():
    try:
        nums = int(entry_students.get())
        if nums <= 0:
            messagebox.showerror("오류", "올바른 학생 수를 입력해주세요!")
            generate_candidate_buttons()
            return False
    except ValueError:
        messagebox.showerror("오류", "올바른 학생 수를 입력해주세요!")
        generate_candidate_buttons()
        return False

    if not add_excluded_numbers():
        generate_candidate_buttons()
        return False

    total_seats = TOTAL_SEATS
    active_seats = total_seats - len(selected)
    available_students = [i for i in range(1, nums + 1) if i not in excluded]
    if active_seats != len(available_students):
        messagebox.showerror("오류", "활성화된 자리 수와 배정할 학생 수가 일치하지 않습니다!")
        generate_candidate_buttons()
        return False
    return True

def start_countdown_and_generate_seats():
    # 자리 생성 단계를 거치지 않았으면 경고
    if not seat_buttons:
        messagebox.showerror("오류", "먼저 자리 생성을 완료해주세요!")
        return
    
    # 이미 자리가 배치되어 있는지 확인 (버튼에 숫자가 있으면 배치된 것)
    is_already_assigned = False
    for row_buttons in seat_buttons:
        for btn in row_buttons:
            if btn.winfo_exists() and btn['text'] not in ['X', '']:
                # 숫자가 있으면 이미 배치된 것
                is_already_assigned = True
                break
        if is_already_assigned:
            break
    
    # 이미 배치되어 있으면 자리 생성 상태로 초기화
    if is_already_assigned:
        generate_candidate_buttons()
        
    if not can_assign_seats():
        return
    # 반복 횟수 확인 및 제한 (빈값은 1회, 0 이하는 오류 처리)
    repeat_str = entry_repeat.get().strip()
    if repeat_str == '':
        repeat_raw = 1
    else:
        try:
            repeat_raw = int(repeat_str)
        except Exception:
            messagebox.showerror("오류", "자동 반복 횟수는 숫자로 입력해주세요!")
            return
        if repeat_raw <= 0:
            messagebox.showerror("오류", "자동 반복 횟수는 1 이상 입력해주세요!")
            return
    if repeat_raw > MAX_REPEAT:
        # 최대 반복 횟수 초과 시 경고만 표시하고 배치를 진행하지 않음
        messagebox.showwarning("경고", f"자동 반복 횟수는 최대 {MAX_REPEAT}회까지 가능합니다.")
        return

    total_runs = repeat_raw
    show_progress = total_runs > 1
    theme = THEMES[current_theme]

    set_inputs_state('disabled')
    
    # 카운트다운 시작 시 기존 책상 버튼들을 비활성화
    for row_buttons in seat_buttons:
        for btn in row_buttons:
            if btn.winfo_exists():
                btn.config(state='disabled')
    
    # 깔끔한 카운트다운: 3 → 2 → 1
    def show_countdown(number):
        if number > 0:
            countdown_label.config(text=str(number), fg=theme['countdown_fg'])
            root.after(600, lambda: show_countdown(number - 1))
        else:
            countdown_label.config(text='')
            start_after_countdown()
    
    def start_after_countdown():
        if total_runs == 1:
            # 1회 실행: 애니메이션과 함께
            generate_seats_with_animation()
            root.after(2000, lambda: set_inputs_state('normal'))
        else:
            # 자동 반복 상태 초기화 및 활성화
            global auto_run_active, speed_factor, speed_key_press_count, speed_key_press_times, boost_count
            auto_run_active = True
            speed_factor = 1.0
            speed_key_press_count = 0
            speed_key_press_times = []
            boost_count = 0
            run_generate_iterations(total_runs, total_runs, show_progress)
    
    # 카운트다운 시작
    show_countdown(3)


def run_generate_iterations(remaining, total, show_progress=True):
    """자리 배치를 remaining 횟수만큼 자동으로 반복 실행"""
    global auto_run_active, speed_factor, speed_key_press_count, speed_key_press_times, boost_count
    theme = THEMES[current_theme]
    
    if remaining <= 0:
        countdown_label.config(text='')
        set_inputs_state('normal')
        auto_run_active = False
        speed_factor = 1.0
        speed_key_press_count = 0
        speed_key_press_times = []
        boost_count = 0
        return
    
    done = total - remaining + 1
    is_last = remaining == 1
    
    if show_progress:
        # 진행 상황 표시 - 심플하게
        speed_text = f" x{2 ** boost_count}" if boost_count > 0 else ""
        percentage = int((done / total) * 100)
        countdown_label.config(
            text=f"{percentage}% ({done}/{total}){speed_text}",
            fg=theme['info_fg']
        )
    
    # 마지막 회차는 긴 애니메이션, 그 외에는 빠른 애니메이션
    if is_last:
        # 완료 메시지
        countdown_label.config(text='완료!', fg=theme['countdown_fg'])
        
        # 마지막 회차 - 완전한 애니메이션 효과
        generate_seats_with_animation(is_last_iteration=True)
        
        # 애니메이션이 완전히 끝난 후 마무리
        def _finalize_after_last():
            root.after(1500, lambda: countdown_label.config(text=''))
            set_inputs_state('normal')
            global auto_run_active, speed_factor, speed_key_press_count, speed_key_press_times, boost_count
            auto_run_active = False
            speed_factor = 1.0
            speed_key_press_count = 0
            speed_key_press_times = []
            boost_count = 0
        # 애니메이션 완료 시간(약 2초) 후 마무리 실행
        root.after(2000, _finalize_after_last)
    else:
        # 중간 회차 - 빠른 애니메이션
        generate_seats_with_quick_animation()
        
        # 자동 실행 중에는 클릭 방지를 위해 버튼 비활성화 유지
        for row_buttons in seat_buttons:
            for btn in row_buttons:
                if btn.winfo_exists() and btn['text'] != 'X':
                    btn.config(state='disabled')
        
        # 반복 횟수에 따라 자동으로 간격 조정
        # total이 10 이하: 기본 속도
        # total이 20: 0.75배 속도 (1.33배 빠름)
        # total이 50: 0.5배 속도 (2배 빠름)
        # total이 100 이상: 0.3배 속도 (3.33배 빠름)
        if total <= 10:
            auto_speed_factor = 1.0
        elif total <= 20:
            auto_speed_factor = 0.75
        elif total <= 50:
            auto_speed_factor = 0.5
        else:
            auto_speed_factor = 0.3
        
        # 현재 속도 배수와 자동 속도 배수 모두 적용
        delay_ms = int(AUTO_RUN_DELAY_MS * speed_factor * auto_speed_factor)
        root.after(delay_ms, lambda: run_generate_iterations(remaining - 1, total, show_progress))

def on_speed_key_press(event=None):
    """
    자동 반복 중 '1' 키를 1초 이내에 2회 입력하면 속도 2배 증가 (최대 x8)
    - 1초 이내에 2번 눌러야 배속 적용
    - 화면에는 속도 배수만 표시
    - 최대 x8 (3회 부스트)까지 가능
    """
    import time
    global speed_key_press_count, speed_key_press_times, speed_factor, boost_count
    theme = THEMES[current_theme]
    
    # 자동 반복 모드가 아니면 무시
    if not auto_run_active:
        return
    
    current_time = time.time()
    
    # 1초 이상 지난 입력 기록 제거
    speed_key_press_times = [t for t in speed_key_press_times if current_time - t <= speed_boost_timeout]
    
    # 현재 입력 시간 추가
    speed_key_press_times.append(current_time)
    speed_key_press_count = len(speed_key_press_times)
    
    # 2번째 입력 시 배속 적용
    if speed_key_press_count >= 2:
        try:
            current_text = countdown_label.cget('text')
            if not current_text:
                return
                
            # 기존 속도 표시 제거
            base_text = current_text
            if 'x' in base_text and base_text.count('x') > 0:
                # 속도 배수 표시 제거
                base_text = base_text.rsplit('x', 1)[0].rstrip()
            
            # 배속 적용
            if boost_count < 3:  # 최대 x8까지만 (부스트 3회)
                boost_count += 1
                speed_factor *= 0.5  # 딜레이 절반 = 속도 2배
                
                # 속도 배수 표시 추가
                speed_multiplier = 2 ** boost_count
                speed_text = f" x{speed_multiplier}"
                countdown_label.config(text=f"{base_text}{speed_text}")
                
                # 시각적 피드백 (깜빡임 효과)
                def blink_effect(times=2):
                    if times <= 0:
                        return
                    current_color = countdown_label.cget('fg')
                    # 색상 토글
                    new_color = 'white' if current_color == theme['countdown_fg'] else theme['countdown_fg']
                    countdown_label.config(fg=new_color)
                    root.after(100, lambda: blink_effect(times - 1))
                
                blink_effect()
            else:
                # 이미 최대 속도 (x8)
                speed_multiplier = 2 ** boost_count
                speed_text = f" x{speed_multiplier} (MAX)"
                countdown_label.config(text=f"{base_text}{speed_text}")
            
            # 입력 기록 초기화
            speed_key_press_times = []
            speed_key_press_count = 0
                
        except Exception as e:
            # 오류 발생 시 디버깅용 출력
            print(f"속도 조절 중 오류: {e}")
            pass

def set_border_to_merged_range(ws, merge_range, border):
    min_col, min_row, max_col, max_row = range_boundaries(merge_range)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border

def open_update_editor():
    """업데이트 내용 보기 창을 엽니다"""
    global update_window
    
    # 이미 열려있다면 포커스만 이동
    if update_window and update_window.winfo_exists():
        update_window.lift()
        update_window.focus_force()
        return
    
    # 새 창 생성
    theme = THEMES[current_theme]
    update_window = Toplevel(root)
    update_window.title("업데이트 내용")
    update_window.geometry("600x600")
    update_window.config(bg=theme['bg'])
    update_window.resizable(True, True)
    
    # 창이 닫힐 때 전역 변수 정리
    def on_closing():
        global update_window
        try:
            update_window.destroy()
        except:
            pass
        update_window = None
    
    update_window.protocol("WM_DELETE_WINDOW", on_closing)
    
    # 메인 프레임
    main_frame = Frame(update_window, bg=theme['bg'])
    main_frame.pack(expand=True, fill='both', padx=20, pady=20)
    
    # 제목
    title_label = Label(main_frame, text="업데이트 내용", 
                       font=('맑은 고딕', 16, 'bold'), bg=theme['bg'], fg=theme['title_fg'])
    title_label.pack(pady=(0, 20))
    
    # 설명
    info_label = Label(main_frame, text="프로그램의 업데이트 내용을 확인할 수 있습니다.", 
                       font=('맑은 고딕', 10), bg=theme['bg'], fg=theme['info_fg'])
    info_label.pack(pady=(0, 10))
    
    # 텍스트 보기 영역
    text_frame = Frame(main_frame, bg=theme['bg'])
    text_frame.pack(expand=True, fill='both', pady=(0, 20))
    
    # 스크롤바가 있는 텍스트 위젯 (읽기 전용)
    text_widget = Text(text_frame, wrap='word', font=('맑은 고딕', 11), 
                      bg=theme['input_bg'], fg=theme['input_fg'], relief='solid', bd=1, state='disabled')
    scrollbar = Scrollbar(text_frame, orient='vertical', command=text_widget.yview)
    text_widget.configure(yscrollcommand=scrollbar.set)
    
    text_widget.pack(side='left', expand=True, fill='both')
    scrollbar.pack(side='right', fill='y')
    
    # 기존 업데이트 내용 로드
    try:
        # 현재 스크립트 위치 기준으로 파일 경로 설정
        import os
        script_dir = os.path.dirname(os.path.abspath(__file__))
        log_file_path = os.path.join(script_dir, 'update_log.txt')
        
        with open(log_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
            text_widget.config(state='normal')
            text_widget.insert('1.0', content)
            text_widget.config(state='disabled')
    except FileNotFoundError:
        text_widget.config(state='normal')
        text_widget.insert('1.0', "업데이트 내용을 불러올 수 없습니다.")
        text_widget.config(state='disabled')
    except Exception as e:
        text_widget.config(state='normal')
        text_widget.insert('1.0', f"파일을 읽는 중 오류가 발생했습니다: {str(e)}")
        text_widget.config(state='disabled')
    
    # 버튼 프레임
    button_frame = Frame(main_frame, bg=theme['bg'])
    button_frame.pack(pady=(0, 10))
    
    # 닫기 버튼
    close_btn = Button(button_frame, text='닫기', 
                      command=on_closing,
                      font=('맑은 고딕', 11, 'bold'), bg=theme['button_excel'], fg=theme['button_fg'],
                      relief='raised', bd=2, width=10)
    close_btn.pack()

def save_update_content(content):
    """업데이트 내용을 파일에 저장합니다"""
    try:
        import os
        script_dir = os.path.dirname(os.path.abspath(__file__))
        log_file_path = os.path.join(script_dir, 'update_log.txt')
        
        with open(log_file_path, 'w', encoding='utf-8') as f:
            f.write(content)
        messagebox.showinfo("성공", "업데이트 내용이 저장되었습니다!")
    except Exception as e:
        messagebox.showerror("오류", f"저장 중 오류가 발생했습니다: {str(e)}")

def load_update_content(text_widget):
    """파일에서 업데이트 내용을 다시 로드합니다"""
    try:
        import os
        script_dir = os.path.dirname(os.path.abspath(__file__))
        log_file_path = os.path.join(script_dir, 'update_log.txt')
        
        with open(log_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
            text_widget.config(state='normal')
            text_widget.delete('1.0', 'end')
            text_widget.insert('1.0', content)
            text_widget.config(state='disabled')
        messagebox.showinfo("성공", "업데이트 내용을 새로고침했습니다!")
    except FileNotFoundError:
        messagebox.showerror("오류", "업데이트 로그 파일을 찾을 수 없습니다.")
    except Exception as e:
        messagebox.showerror("오류", f"로드 중 오류가 발생했습니다: {str(e)}")

# 메인 윈도우 생성
root = Tk()
root.title("교실 자리 배치 프로그램")
root.geometry("1000x800")  # 기본 크기 설정
root.config(bg=THEMES[current_theme]['bg'])

# 전체화면에서 중앙 정렬을 위한 메인 컨테이너 프레임
main_container = Frame(root, bg=THEMES[current_theme]['bg'])
main_container.pack(expand=True, fill='both', padx=20, pady=20)

# 타이틀 라벨
title_label = Label(main_container, text=Title_Text, 
                   bg=THEMES[current_theme]['bg'], fg=THEMES[current_theme]['title_fg'], font=('맑은 고딕', 24, 'bold'))
title_label.pack(pady=(0, 15))

# 입력 프레임 생성
input_frame = Frame(main_container, bg=THEMES[current_theme]['bg'])
input_frame.pack(pady=(0, 20))

# 입력 필드들 - 첫 번째 행
label_grade_frame = Frame(input_frame, bg=THEMES[current_theme]['bg'])
label_grade_frame.grid(row=0, column=0, padx=10, pady=5, sticky='e')
Label(label_grade_frame, text='학년', bg=THEMES[current_theme]['bg'], fg=THEMES[current_theme]['fg'], font=('맑은 고딕', 12, 'bold')).pack(side='left')
Label(label_grade_frame, text='*', bg=THEMES[current_theme]['bg'], fg='red', font=('맑은 고딕', 12, 'bold')).pack(side='left')
entry_grade = Entry(input_frame, width=15, font=('맑은 고딕', 12), bd=1, relief='solid', bg=THEMES[current_theme]['input_bg'], fg=THEMES[current_theme]['input_fg'])
entry_grade.grid(row=0, column=1, padx=10, pady=5)

label_group_frame = Frame(input_frame, bg=THEMES[current_theme]['bg'])
label_group_frame.grid(row=0, column=2, padx=10, pady=5, sticky='e')
Label(label_group_frame, text='반', bg=THEMES[current_theme]['bg'], fg=THEMES[current_theme]['fg'], font=('맑은 고딕', 12, 'bold')).pack(side='left')
Label(label_group_frame, text='*', bg=THEMES[current_theme]['bg'], fg='red', font=('맑은 고딕', 12, 'bold')).pack(side='left')
entry_group = Entry(input_frame, width=15, font=('맑은 고딕', 12), bd=1, relief='solid', bg=THEMES[current_theme]['input_bg'], fg=THEMES[current_theme]['input_fg'])
entry_group.grid(row=0, column=3, padx=10, pady=5)

# 두 번째 행
label_students_frame = Frame(input_frame, bg=THEMES[current_theme]['bg'])
label_students_frame.grid(row=1, column=0, padx=10, pady=5, sticky='e')
Label(label_students_frame, text='학생 수\n(1~18)', bg=THEMES[current_theme]['bg'], fg=THEMES[current_theme]['fg'], font=('맑은 고딕', 12, 'bold')).pack(side='left')
Label(label_students_frame, text='*', bg=THEMES[current_theme]['bg'], fg='red', font=('맑은 고딕', 12, 'bold')).pack(side='left')
entry_students = Entry(input_frame, width=15, font=('맑은 고딕', 12), bd=1, relief='solid', bg=THEMES[current_theme]['input_bg'], fg=THEMES[current_theme]['input_fg'])
entry_students.grid(row=1, column=1, padx=10, pady=5)

label_teacher_frame = Frame(input_frame, bg=THEMES[current_theme]['bg'])
label_teacher_frame.grid(row=1, column=2, padx=10, pady=5, sticky='e')
Label(label_teacher_frame, text='담임선생님', bg=THEMES[current_theme]['bg'], fg=THEMES[current_theme]['fg'], font=('맑은 고딕', 12, 'bold')).pack(side='left')
Label(label_teacher_frame, text='*', bg=THEMES[current_theme]['bg'], fg='red', font=('맑은 고딕', 12, 'bold')).pack(side='left')
entry_teacher = Entry(input_frame, width=15, font=('맑은 고딕', 12), bd=1, relief='solid', bg=THEMES[current_theme]['input_bg'], fg=THEMES[current_theme]['input_fg'])
entry_teacher.grid(row=1, column=3, padx=10, pady=5)

# 세 번째 행
label_repeat = Label(input_frame, text='자동 반복 횟수', bg=THEMES[current_theme]['bg'], fg=THEMES[current_theme]['fg'], font=('맑은 고딕', 12, 'bold'))
label_repeat.grid(row=2, column=0, padx=10, pady=5, sticky='e')
entry_repeat = Entry(input_frame, width=15, font=('맑은 고딕', 12), bd=1, relief='solid', bg=THEMES[current_theme]['input_bg'], fg=THEMES[current_theme]['input_fg'])
entry_repeat.grid(row=2, column=1, padx=10, pady=5)

label_exclude = Label(input_frame, text='제외할 번호\n(쉼표로 구분)', bg=THEMES[current_theme]['bg'], fg=THEMES[current_theme]['fg'], font=('맑은 고딕', 12, 'bold'))
label_exclude.grid(row=2, column=2, padx=10, pady=5, sticky='e')
entry_exclude = Entry(input_frame, width=15, font=('맑은 고딕', 12), bd=1, relief='solid', bg=THEMES[current_theme]['input_bg'], fg=THEMES[current_theme]['input_fg'])
entry_exclude.grid(row=2, column=3, padx=10, pady=5)

# 설명 라벨 - 네 번째 행
info_label = Label(input_frame, text="사용법: 1. 정보 입력 → 2. 자리 생성 → 3. 비활성화할 자리 선택 → 4. 자리 배치 → 5. 엑셀 생성", 
                  bg=THEMES[current_theme]['bg'], fg=THEMES[current_theme]['info_fg'], font=('맑은 고딕', 20))
info_label.grid(row=3, column=0, columnspan=4, pady=10)

# 버튼들 - 다섯 번째 행
btn_frame = Frame(input_frame, bg=THEMES[current_theme]['bg'])
btn_frame.grid(row=4, column=0, columnspan=4, pady=10)

btn_generate_candidates = Button(btn_frame, text='자리 생성', 
                               command=generate_candidate_buttons,
                               font=('맑은 고딕', 11, 'bold'), bg=THEMES[current_theme]['button_generate'], fg=THEMES[current_theme]['button_fg'],
                               relief='raised', bd=2, width=10)
btn_generate_candidates.grid(row=0, column=0, padx=5, pady=5)

btn_generate_seats = Button(btn_frame, text='자리 배치', 
                          command=start_countdown_and_generate_seats,
                          font=('맑은 고딕', 11, 'bold'), bg=THEMES[current_theme]['button_assign'], fg=THEMES[current_theme]['button_fg'],
                          relief='raised', bd=2, width=10, state='disabled')
btn_generate_seats.grid(row=0, column=1, padx=5, pady=5)

btn_create_excel = Button(btn_frame, text='엑셀 생성', 
                         command=create_excel_file,
                         font=('맑은 고딕', 11, 'bold'), bg=THEMES[current_theme]['button_excel'], fg=THEMES[current_theme]['button_fg'],
                         relief='raised', bd=2, width=10)
btn_create_excel.grid(row=0, column=2, padx=5, pady=5)

# 입력 필드와 버튼을 리스트로 관리
all_inputs = [
    entry_grade, entry_group, entry_students, entry_teacher, entry_exclude, entry_repeat,
    btn_generate_candidates, btn_generate_seats, btn_create_excel
]

def set_inputs_state(state):
    for widget in all_inputs:
        widget.config(state=state)
    
    # 자리 배치 버튼은 자리 생성이 완료된 후에만 활성화
    if state == 'normal' and not seat_buttons:
        btn_generate_seats.config(state='disabled')

# 칠판 위치 표시 라벨
blackboard_label = Button(input_frame, text="칠판", 
                        font=('맑은 고딕', 11, 'bold'), bg=THEMES[current_theme]['blackboard_bg'], fg=THEMES[current_theme]['blackboard_fg'],
                         relief='raised', bd=2, width=100)
blackboard_label.grid(row=5, column=0, columnspan=4, pady=5)

# 자리 배치 프레임
frame = Frame(main_container, bg=THEMES[current_theme]['bg'])
frame.pack(pady=20)

# 카운트다운 라벨 추가
countdown_label = Label(main_container, text='', font=('맑은 고딕', 40, 'bold'), bg=THEMES[current_theme]['bg'], fg=THEMES[current_theme]['countdown_fg'])
countdown_label.pack(pady=10)

# 왼쪽 하단에 업데이트 내용 버튼 배치
update_button_frame = Frame(main_container, bg=THEMES[current_theme]['bg'])
update_button_frame.pack(side='bottom', anchor='sw', padx=10, pady=5)

update_content_btn = Button(update_button_frame, text='업데이트 내용', 
                           command=open_update_editor,
                           font=('맑은 고딕', 11, 'bold'), bg=THEMES[current_theme]['button_generate'], fg=THEMES[current_theme]['button_fg'],
                           relief='raised', bd=2, width=10)
update_content_btn.pack()

# 테마 변경 버튼 추가 (왼쪽 하단, 업데이트 내용 버튼 위)
theme_button_frame = Frame(main_container, bg=THEMES[current_theme]['bg'])
theme_button_frame.pack(side='bottom', anchor='sw', padx=10, pady=5)

theme_btn = Button(theme_button_frame, text='🎨 테마 변경', 
                  command=open_theme_selector,
                  font=('맑은 고딕', 11, 'bold'), bg=THEMES[current_theme]['button_excel'], fg=THEMES[current_theme]['button_fg'],
                  highlightbackground=THEMES[current_theme]['button_excel'],
                  activebackground=THEMES[current_theme]['button_excel'],
                  relief='raised', bd=2, width=10)
theme_btn.pack()

# 확대/축소 상태 표시 라벨 (root에 직접 배치하여 오른쪽 맨 밑에 배치)
zoom_status_label = Label(root, text="확대/축소: 100%", font=('맑은 고딕', 10), bg=THEMES[current_theme]['bg'], fg=THEMES[current_theme]['fg'])
zoom_status_label.place(relx=1.0, rely=1.0, anchor='se', x=-10, y=-30)

# 단축키 안내 라벨 (root에 직접 배치하여 오른쪽 맨 밑에 배치, zoom_status_label 아래)
shortcut_label = Label(root, text="단축키: ⌘+ 또는 ⌘= (확대) | ⌘- (축소) | ⌘0 (원래 크기)", 
                      font=('맑은 고딕', 9), bg=THEMES[current_theme]['bg'], fg=THEMES[current_theme]['info_fg'])
shortcut_label.place(relx=1.0, rely=1.0, anchor='se', x=-10, y=-10)

# 키보드 단축키 바인딩 (macOS 호환성 향상)
root.bind('<Command-plus>', zoom_in)
root.bind('<Command-equal>', zoom_in)  # Command + = (macOS에서 +와 =이 같은 키)
root.bind('<Command-minus>', zoom_out)
root.bind('<Command-0>', zoom_reset)
root.bind('<Command-Key-0>', zoom_reset)  # macOS 호환성
# 속도 증가 트리거(1 키 4연타) - 포커스에 상관없이 동작하도록 전체 바인딩
root.bind_all('<KeyPress-1>', on_speed_key_press)
# 숫자 키패드의 1도 인식 (필요 시)
root.bind_all('<KeyPress-KP_1>', on_speed_key_press)

# 시작 시 기본 테마 적용 (중복 체크를 우회하기 위해 임시로 다른 값으로 설정)
current_theme = None
apply_theme('green')

root.mainloop()